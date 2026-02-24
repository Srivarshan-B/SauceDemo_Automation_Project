from openpyxl import load_workbook


class DataReader:

    def _init_(self, file_path):
        self.wb = load_workbook(file_path)
        self.sheet = self.wb.active

    def get_data(self):
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data